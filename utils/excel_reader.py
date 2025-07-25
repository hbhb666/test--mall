import os
from openpyxl import load_workbook
from config.config import base_url  # 从配置文件导入基础URL


def read_excel_test_cases(excel_path: str, sheet_name: str) -> list:
    """
    读取Excel中的测试用例，转换为字典列表。
    :param excel_path: Excel文件路径（相对/绝对路径）
    :param sheet_name:  sheet名称
    :return: 测试用例列表（每个元素是字典，键为列名）
    """
    # 1. 校验Excel文件是否存在
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel文件不存在：{excel_path}")


    # 2. 加载Excel workbook
    workbook=load_workbook(excel_path)
    sheet=workbook[sheet_name]
    if not sheet:
        raise ValueError(f"Sheet不存在：{sheet_name}")

    # 3. 提取列名（第一行）
    columns = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    # 4. 遍历行，转换为字典
    test_cases=[]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        case=dict(zip(columns, row))

    # 5. 替换变量（如{{portal.mall}}→base_url）
        if "{{portal.mall}}" in case.get("接口地址", ""):
            case["接口地址"] = case["接口地址"].replace("{{portal.mall}}", base_url)

        # 6. 处理请求头（字符串→字典）
        if case.get("请求头"):
            headers={}
            for header in case["请求头"].split(";"):
                if":"in header:
                    key, value = header.split(":",1)
                    headers[key.strip()] = value.strip()
            case["请求头"] = headers
    # 7. 处理参数输入（根据请求方式转换格式）
        request_method = case.get("请求方式", "").upper()
        case["请求方式"] = request_method  # 回写规范化后的请求方式
        params_input = case.get("参数输入", "")
        # 去除参数输入中的前后空格和换行符，并格式化为有效的JSON字符串
        if isinstance(params_input, str):
            # 先去除首尾空白
            params_input = params_input.strip()
            # 替换换行符和多余空格为单个空格，避免JSON解析错误
            params_input = params_input.replace('\n', ' ').replace('\r', ' ').replace('  ', ' ')
            case["参数输入"] = params_input
    # 8. 添加到测试用例列表
        test_cases.append(case)

    # 9. 关闭workbook
    workbook.close()
    # for case1 in test_cases:
    #     print(case1)
    # print(f"成功读取{len(test_cases)}条测试用例")
    return test_cases


if __name__ == '__main__':
    test_cases = read_excel_test_cases("../data/mall测试用例.xlsx", "Sheet1")
